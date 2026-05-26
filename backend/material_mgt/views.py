import os
import requests
import logging
from uuid import UUID
from django.utils import timezone
from django.db.models import Avg, Count, Prefetch, Q
from django.utils.translation import gettext_lazy as _
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
from rest_framework import status
from rest_framework.exceptions import ValidationError
from rest_framework import viewsets
from rest_framework.viewsets import ModelViewSet
from rest_framework.permissions import BasePermission, IsAuthenticated, SAFE_METHODS
from rest_framework.decorators import action
from rest_framework import filters
from django_filters.rest_framework import DjangoFilterBackend
from google import genai

from user_mgt.access import (
    get_user_library,
    is_super_admin,
    normalize_role,
    has_global_material_access,
)
from .models import (
    PhysicalMaterial,
    DigitalMaterial,
    MaterialFeedback,
    MaterialFavorite,
    MaterialBookmark,
    MaterialTransferRequest,
)
from .serializers import (
    PhysicalMaterialSerializer,
    DigitalMaterialSerializer,
    MaterialFeedbackSerializer,
    MaterialFavoriteSerializer,
    MaterialBookmarkSerializer,
    MaterialTransferRequestSerializer,
)
from transactions.models import Borrow, Reservation
from .services import generate_material_description, lookup_book_metadata

logger = logging.getLogger(__name__)

def _get_gemini_model(env_name: str, default: str = "gemini-2.0-flash") -> str:
    return os.getenv(env_name, default).strip() or default


def _request_gemini_text(*, model: str, prompt: str, max_output_tokens: int = 400):
    api_key = os.getenv("GEMINI_API_KEY", "").strip()

    if not api_key:
        return {
            "success": False,
            "status": status.HTTP_503_SERVICE_UNAVAILABLE,
            "error": "Gemini API key not configured.",
            "text": "",
        }

    try:
        client = genai.Client(api_key=api_key)

        response = client.models.generate_content(
            model=model,
            contents=prompt,
        )

        text = response.text.strip() if response.text else ""

        return {
            "success": True,
            "status": status.HTTP_200_OK,
            "error": "",
            "text": text,
        }

    except Exception as e:
        return {
            "success": False,
            "status": status.HTTP_502_BAD_GATEWAY,
            "error": str(e),
            "text": "",
        }

def _parse_bool(value, default=False):
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _resolve_material_or_error(material_type, material_id):
    material_type_norm = str(material_type or "").strip().lower()
    material_id_str = str(material_id or "").strip()

    try:
        material_uuid = UUID(material_id_str)
    except (TypeError, ValueError):
        raise ValidationError({"material_id": _("Invalid material id.")})

    if material_type_norm == "physical":
        material = PhysicalMaterial.objects.filter(pk=material_uuid).first()
        if not material:
            raise ValidationError({"material_id": "Physical material not found."})
        return material_type_norm, material

    if material_type_norm == "digital":
        material = DigitalMaterial.objects.filter(pk=material_uuid).first()
        if not material:
            raise ValidationError({"material_id": "Digital material not found."})
        return material_type_norm, material

    raise ValidationError({"material_type": "Use either 'physical' or 'digital'."})


def _user_can_access_material(user, material):
    if has_global_material_access(user):
        return True
    user_library = get_user_library(user)
    return bool(user_library and getattr(material, "library_id", None) == user_library.id)


class IsOwnerOrReadOnly(BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated)

    def has_object_permission(self, request, view, obj):
        if request.method in SAFE_METHODS:
            return True
        return obj.user_id == request.user.id


def _get_staff_profile_or_error(user):
    staff_roles = {"STACKSTAFF", "TECHNICALSTAFF", "FRONTDESKSTAFF", "ADMIN", "SUPERADMIN"}
    if normalize_role(getattr(user, "role", None)) not in staff_roles:
        raise ValidationError({"detail": "Only staff-role users can create materials."})
    return user

class PhysicalMaterialViewSet(ModelViewSet):
    queryset = PhysicalMaterial.objects.select_related("library", "created_by").prefetch_related(
        Prefetch("feedbacks", queryset=MaterialFeedback.objects.select_related("user").order_by("-updated_at"))
    ).annotate(
        average_rating=Avg("feedbacks__rating"),
        ratings_count=Count("feedbacks"),
    ).order_by("title", "id")
    serializer_class = PhysicalMaterialSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'genre', 'language', 'department', 'condition', 'location']
    search_fields = ['title', 'author', 'isbn', 'barcode']
    ordering_fields = ['title', 'published_date', 'average_rating', 'ratings_count', 'id']
    # permission_classes = [IsAuthenticated, IsTechnicalStaffForWrite]

    def get_queryset(self):
        queryset = super().get_queryset()

        if has_global_material_access(self.request.user):
            return queryset

        # If location is explicitly requested via query params, skip role-based filtering
        requested_location = self.request.query_params.get('location', '').strip()

        if not requested_location:
            role = normalize_role(getattr(self.request.user, "role", None))
            if role == "FRONTDESKSTAFF":
                queryset = queryset.filter(location="SHELF")
            elif role == "STACKSTAFF":
                queryset = queryset.filter(location="STACK")
            elif role in {"TECHNICALSTAFF", "ADMIN", "SUPERADMIN", "DEPARTMENTHEAD"}:
                pass
            elif role == "MEMBER":
                queryset = queryset.exclude(location__in=["SHELF", "STACK"])
            elif self.request.user.is_authenticated:
                queryset = queryset.exclude(location__in=["SHELF", "STACK"])
            # Anonymous users can browse available materials without library restriction.

        actor_library = get_user_library(self.request.user)
        if actor_library:
            queryset = queryset.filter(library=actor_library)

        return queryset

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        staff_profile = _get_staff_profile_or_error(request.user)
        actor_library = get_user_library(request.user)
        if not actor_library and not is_super_admin(request.user):
            raise ValidationError({"library": "Your account is not assigned to a library yet."})

        data = serializer.validated_data
        total_copies = int(data.get("total_copies") or 0)
        if total_copies < 1:
            raise ValidationError({"total_copies": "total_copies must be at least 1."})

        available_copies = data.get("available_copies")
        if available_copies is None:
            available_copies = total_copies
        elif int(available_copies) > total_copies:
            raise ValidationError(
                {"available_copies": "available_copies cannot exceed total_copies."}
            )

        material = serializer.save(
            created_by=staff_profile,
            available_copies=available_copies,
            library=data.get("library") if is_super_admin(request.user) else actor_library,
        )
        # Optionally auto-generate description if requested and empty
        try:
            generate_flag = _parse_bool(request.data.get("generate_description"), default=False) or _parse_bool(request.query_params.get("generate_description"), default=False)
            if generate_flag and not getattr(material, "description", None):
                desc, used_model = generate_material_description(material.title, material.author)
                if desc:
                    material.description = desc
                    material.save(update_fields=["description"])
        except Exception:
            # Do not block creation on generation errors
            logger.exception("Failed to auto-generate description for physical material %s", material.pk)
        output = self.get_serializer(material)
        return Response(output.data, status=status.HTTP_201_CREATED)

    def perform_update(self, serializer):
        actor_library = get_user_library(self.request.user)
        if not actor_library and not is_super_admin(self.request.user):
            raise ValidationError({"library": "Your account is not assigned to a library yet."})
        serializer.save(
            library=serializer.validated_data.get("library") if is_super_admin(self.request.user) else actor_library
        )

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        if not _user_can_access_material(request.user, instance):
            raise ValidationError({"detail": "You can only access materials in your library."})

        if not getattr(instance, "description", None):
            try:
                desc, used_model = generate_material_description(instance.title, instance.author)
                if desc:
                    instance.description = desc
                    instance.save(update_fields=["description"])
            except Exception:
                logger.exception("Failed to generate description on retrieve for physical material %s", instance.pk)

        serializer = self.get_serializer(instance)
        return Response(serializer.data)


class DigitalMaterialViewSet(ModelViewSet):
    queryset = DigitalMaterial.objects.select_related("library", "created_by").prefetch_related(
        Prefetch("feedbacks", queryset=MaterialFeedback.objects.select_related("user").order_by("-updated_at"))
    ).annotate(
        average_rating=Avg("feedbacks__rating"),
        ratings_count=Count("feedbacks"),
    ).order_by("title", "id")
    serializer_class = DigitalMaterialSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'genre', 'language', 'department', 'format']
    search_fields = ['title', 'author', 'isbn']
    ordering_fields = ['title', 'published_date', 'average_rating', 'ratings_count', 'id']
    # permission_classes = [IsAuthenticated, IsTechnicalStaffForWrite]

    def get_queryset(self):
        queryset = super().get_queryset()
        if has_global_material_access(self.request.user):
            return queryset

        actor_library = get_user_library(self.request.user)
        if not actor_library:
            return queryset.none()
        return queryset.filter(library=actor_library)

    def perform_create(self, serializer):
        actor_library = get_user_library(self.request.user)
        if not actor_library and not is_super_admin(self.request.user):
            raise ValidationError({"library": "Your account is not assigned to a library yet."})
        instance = serializer.save(
            created_by=_get_staff_profile_or_error(self.request.user),
            library=serializer.validated_data.get("library") if is_super_admin(self.request.user) else actor_library,
        )
        if not instance.cover_image and generate_pdf_cover_image(instance):
            instance.save(update_fields=["cover_image", "cover_generated_at"])
        # Optionally auto-generate description if requested and empty
        try:
            generate_flag = _parse_bool(self.request.data.get("generate_description"), default=False) or _parse_bool(self.request.query_params.get("generate_description"), default=False)
            if generate_flag and not getattr(instance, "description", None):
                desc, used_model = generate_material_description(instance.title, instance.author)
                if desc:
                    instance.description = desc
                    instance.save(update_fields=["description"])
        except Exception:
            logger.exception("Failed to auto-generate description for digital material %s", instance.pk)

    def perform_update(self, serializer):
        actor_library = get_user_library(self.request.user)
        if not actor_library and not is_super_admin(self.request.user):
            raise ValidationError({"library": "Your account is not assigned to a library yet."})
        instance = serializer.save(
            library=serializer.validated_data.get("library") if is_super_admin(self.request.user) else actor_library
        )
        if not instance.cover_image and generate_pdf_cover_image(instance):
            instance.save(update_fields=["cover_image", "cover_generated_at"])
        
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        if not _user_can_access_material(request.user, instance):
            raise ValidationError({"detail": "You can only access materials in your library."})

        if not getattr(instance, "description", None):
            try:
                desc, used_model = generate_material_description(instance.title, instance.author)
                if desc:
                    instance.description = desc
                    instance.save(update_fields=["description"])
            except Exception:
                logger.exception("Failed to generate description on retrieve for digital material %s", instance.pk)

        serializer = self.get_serializer(instance)
        return Response(serializer.data)


class MaterialFeedbackViewSet(ModelViewSet):
    serializer_class = MaterialFeedbackSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrReadOnly]
    queryset = MaterialFeedback.objects.select_related("user", "physical_material", "digital_material")

    def get_queryset(self):
        queryset = self.queryset

        material_type = self.request.query_params.get("material_type")
        material_id = self.request.query_params.get("material_id")
        mine_only = _parse_bool(self.request.query_params.get("mine"), default=False)

        if mine_only:
            queryset = queryset.filter(user=self.request.user)

        if bool(material_type) ^ bool(material_id):
            raise ValidationError(
                {"detail": "Provide material_type and material_id together."}
            )

        if material_type and material_id:
            material_type_norm = material_type.strip().lower()
            if material_type_norm == "physical":
                queryset = queryset.filter(physical_material_id=material_id)
            elif material_type_norm == "digital":
                queryset = queryset.filter(digital_material_id=material_id)
            else:
                raise ValidationError({"material_type": "Use either 'physical' or 'digital'."})

        return queryset

    def perform_create(self, serializer):
        target_material = serializer.validated_data.get("physical_material") or serializer.validated_data.get("digital_material")
        if target_material and not _user_can_access_material(self.request.user, target_material):
            raise ValidationError({"detail": "You can only comment on materials from your library."})
        serializer.save(user=self.request.user)

    @action(detail=False, methods=["get"], url_path=r"by-material/(?P<material_type>[^/.]+)/(?P<material_id>[^/.]+)")
    def by_material(self, request, material_type=None, material_id=None):
        if not material_type or not material_id:
            raise ValidationError({"detail": "material_type and material_id are required."})

        material_type_norm, material = _resolve_material_or_error(material_type, material_id)
        if material_type_norm == "physical":
            queryset = self.queryset.filter(physical_material=material)
        else:
            queryset = self.queryset.filter(digital_material=material)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class MaterialFavoriteViewSet(ModelViewSet):
    serializer_class = MaterialFavoriteSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrReadOnly]
    queryset = MaterialFavorite.objects.select_related("user", "physical_material", "digital_material")

    def get_queryset(self):
        queryset = self.queryset.filter(user=self.request.user)

        material_type = self.request.query_params.get("material_type")
        material_id = self.request.query_params.get("material_id")
        if bool(material_type) ^ bool(material_id):
            raise ValidationError(
                {"detail": "Provide material_type and material_id together."}
            )
        if material_type and material_id:
            material_type_norm = material_type.strip().lower()
            if material_type_norm == "physical":
                queryset = queryset.filter(physical_material_id=material_id)
            elif material_type_norm == "digital":
                queryset = queryset.filter(digital_material_id=material_id)
            else:
                raise ValidationError({"material_type": "Use either 'physical' or 'digital'."})

        return queryset

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        physical_material = serializer.validated_data.get("physical_material")
        digital_material = serializer.validated_data.get("digital_material")
        target_material = physical_material or digital_material
        if target_material and not _user_can_access_material(request.user, target_material):
            raise ValidationError({"detail": "You can only favorite materials from your library."})

        favorite, created = MaterialFavorite.objects.get_or_create(
            user=request.user,
            physical_material=physical_material,
            digital_material=digital_material,
        )

        output = self.get_serializer(favorite)
        return Response(
            output.data,
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


class MaterialBookmarkViewSet(ModelViewSet):
    serializer_class = MaterialBookmarkSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrReadOnly]
    queryset = MaterialBookmark.objects.select_related("user", "physical_material", "digital_material")

    def get_queryset(self):
        queryset = self.queryset.filter(user=self.request.user)

        material_type = self.request.query_params.get("material_type")
        material_id = self.request.query_params.get("material_id")
        if bool(material_type) ^ bool(material_id):
            raise ValidationError(
                {"detail": "Provide material_type and material_id together."}
            )
        if material_type and material_id:
            material_type_norm = material_type.strip().lower()
            if material_type_norm == "physical":
                queryset = queryset.filter(physical_material_id=material_id)
            elif material_type_norm == "digital":
                queryset = queryset.filter(digital_material_id=material_id)
            else:
                raise ValidationError({"material_type": "Use either 'physical' or 'digital'."})

        return queryset

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        physical_material = serializer.validated_data.get("physical_material")
        digital_material = serializer.validated_data.get("digital_material")
        target_material = physical_material or digital_material
        if target_material and not _user_can_access_material(request.user, target_material):
            raise ValidationError({"detail": "You can only bookmark materials from your library."})

        bookmark, created = MaterialBookmark.objects.get_or_create(
            user=request.user,
            physical_material=physical_material,
            digital_material=digital_material,
        )

        output = self.get_serializer(bookmark)
        return Response(
            output.data,
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


class MaterialInteractionStatsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        material_type = request.query_params.get("material_type")
        material_id = request.query_params.get("material_id")

        if not material_type or not material_id:
            raise ValidationError({"detail": "material_type and material_id are required."})

        material_type_norm, material = _resolve_material_or_error(material_type, material_id)
        if not _user_can_access_material(request.user, material):
            raise ValidationError({"detail": "You can only access interaction stats for materials in your library."})

        feedback_qs = MaterialFeedback.objects.all()
        favorites_qs = MaterialFavorite.objects.all()
        bookmarks_qs = MaterialBookmark.objects.all()

        if material_type_norm == "physical":
            feedback_qs = feedback_qs.filter(physical_material=material)
            favorites_qs = favorites_qs.filter(physical_material=material)
            bookmarks_qs = bookmarks_qs.filter(physical_material=material)
            mine_filter = {"physical_material": material, "user": request.user}
        else:
            feedback_qs = feedback_qs.filter(digital_material=material)
            favorites_qs = favorites_qs.filter(digital_material=material)
            bookmarks_qs = bookmarks_qs.filter(digital_material=material)
            mine_filter = {"digital_material": material, "user": request.user}

        aggregate = feedback_qs.aggregate(
            average_rating=Avg("rating"),
            ratings_count=Count("id"),
            comments_count=Count("id", filter=~Q(comment="")),
        )

        my_feedback = MaterialFeedback.objects.filter(**mine_filter).first()

        return Response(
            {
                "material_type": material_type_norm,
                "material_id": str(material.id),
                "material_title": material.title,
                "average_rating": round(float(aggregate["average_rating"] or 0), 2),
                "ratings_count": int(aggregate["ratings_count"] or 0),
                "comments_count": int(aggregate["comments_count"] or 0),
                "favorites_count": favorites_qs.count(),
                "bookmarks_count": bookmarks_qs.count(),
                "is_favorited": MaterialFavorite.objects.filter(**mine_filter).exists(),
                "is_bookmarked": MaterialBookmark.objects.filter(**mine_filter).exists(),
                "my_rating": my_feedback.rating if my_feedback else None,
                "my_comment": my_feedback.comment if my_feedback else "",
            },
            status=status.HTTP_200_OK,
        )





def _build_chat_prompt(prompt, history=None):
    sanitized_history = history if isinstance(history, list) else []
    lines = [
        "You are a helpful virtual assistant for a university digital library system.",
        "Answer clearly and politely.",
        "Focus on library policies, reservations, borrowing, returns, fines, and material discovery.",
        "If you are unsure, say so and avoid inventing facts.",
        "",
        "Conversation so far:",
    ]

    for item in sanitized_history[-10:]:
        role = str(item.get("role", "")).strip().lower()
        content = str(item.get("content", "")).strip()
        if not content:
            continue
        speaker = "User" if role == "user" else "Assistant"
        lines.append(f"{speaker}: {content}")

    lines.append(f"User: {prompt}")
    lines.append("Assistant:")
    return "\n".join(lines)


def _build_library_context(user):
    if not user or not user.is_authenticated:
        return {}

    borrows = Borrow.objects.select_related("material").filter(member=user).exclude(status="RETURNED").order_by("due_date")[:5]
    reservations = Reservation.objects.select_related("material_id").filter(member=user, status="RESERVED").order_by("reserve_date")[:5]
    available_books = PhysicalMaterial.objects.filter(available_copies__gt=0).order_by("title")[:10]

    return {
        "borrowed_materials": [
            {"title": row.material.title, "due_date": row.due_date.date().isoformat(), "status": row.status}
            for row in borrows
        ],
        "reservations": [
            {"title": row.material_id.title, "expiry_date": row.expiry_date.date().isoformat(), "status": row.status}
            for row in reservations
        ],
        "available_books": [row.title for row in available_books],
    }


class LibraryAssistantChatAPIView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request, *args, **kwargs):
        prompt = str(request.data.get("prompt", "")).strip()
        history = request.data.get("history", [])
        language = str(request.data.get("language", "en")).strip().lower()
        
        if not prompt:
            raise ValidationError({"prompt": _("Prompt is required.")})
        
        # Try multiple models in order of preference
        models_to_try = [
            _get_gemini_model("GEMINI_CHAT_MODEL", default="gemini-2.0-flash"),
            "gemini-2.5-flash",  # Fallback 1
            "gemini-1.5-pro",    # Fallback 2 (if still available)
        ]
        
        context_prefix = (
            f"Preferred response language: {'Amharic' if language.startswith('am') else 'English'}.\n"
            f"Library data context: {_build_library_context(request.user)}\n"
        )
        
        last_error = None
        for model in models_to_try:
            result = _request_gemini_text(
                model=model,
                prompt=context_prefix + _build_chat_prompt(prompt, history),
                max_output_tokens=400,
            )
            
            if result["success"]:
                message = result["text"] or _("I am sorry, I could not generate an answer right now.")
                return Response({"message": message, "model": model}, status=status.HTTP_200_OK)
            
            last_error = result
            # Don't retry quota errors immediately
            if "429" in str(result.get("error", "")):
                break
        
        # Provide helpful error message
        if "quota" in str(last_error.get("error", "")).lower():
            return Response(
                {"detail": "API quota exceeded. Please try again later or contact support."},
                status=status.HTTP_429_TOO_MANY_REQUESTS
            )
        
        return Response({"detail": last_error.get("error", "Unknown error")}, 
                       status=last_error.get("status", 500))

class GenerateMaterialDescriptionAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        title = str(request.data.get("title", "")).strip()
        author = str(request.data.get("author", "")).strip()
        material_type = request.data.get("material_type")
        material_id = request.data.get("material_id")

        if not title or not author:
            raise ValidationError({"detail": "Both title and author are required."})

        models_to_try = [
            _get_gemini_model("GEMINI_DESCRIPTION_MODEL", default="gemini-2.0-flash"),
            "gemini-2.5-flash",
            "gemini-2.0-flash-lite",
            "gemini-1.5-pro",
        ]
        seen = set()
        models_to_try = [m for m in models_to_try if m and not (m in seen or seen.add(m))]

        prompt = (
            f"Title: {title}\n"
            f"Author: {author}\n\n"
            "Write a concise, helpful library catalog description (80-130 words). "
            "Use clear, neutral language and avoid inventing specific facts."
        )

        last_error = None
        used_model = models_to_try[0]
        description = ""

        for model in models_to_try:
            used_model = model
            result = _request_gemini_text(
                model=model,
                prompt=prompt,
                max_output_tokens=220,
            )
            if result["success"] and result["text"]:
                description = result["text"]
                break
            last_error = result

        if not description:
            if last_error and not last_error.get("success"):
                return Response({"detail": last_error.get("error", "Unknown error")}, status=last_error.get("status", 502))
            description = f"{title} by {author} is a library material available for borrowing."

        saved = False
        if material_type and material_id:
            try:
                mtype_norm, material = _resolve_material_or_error(material_type, material_id)
                # ensure user can access and update
                if not _user_can_access_material(request.user, material):
                    raise ValidationError({"detail": "You cannot modify materials outside your library."})
                material.description = description
                material.save(update_fields=["description"])
                saved = True
            except ValidationError:
                # If resolving or permission failed, return a validation error
                raise
            except Exception:
                # on unexpected failure, continue but indicate not saved
                saved = False

        return Response(
            {
                "description": description,
                "model": used_model,
                "saved_to_material": bool(saved),
                "material_type": material_type,
                "material_id": material_id,
            },
            status=status.HTTP_200_OK,
        )



def _notify_role_users(*, roles, message, library=None, exclude_user=None):
    from user_mgt.models import Notification, User

    role_keys = {normalize_role(role) for role in roles}
    users = User.objects.filter(status="ACTIVE")

    for user in users:
        if normalize_role(getattr(user, "role", None)) not in role_keys:
            continue
        if exclude_user and user.pk == exclude_user.pk:
            continue
        if library and getattr(user, "library_id", None) != getattr(library, "id", None):
            continue
        Notification.objects.create(
            member_id=user,
            message=str(message)[:200],
            status="UNREAD",
        )


class MaterialTransferRequestViewSet(viewsets.ModelViewSet):
    queryset = MaterialTransferRequest.objects.all().select_related("material", "requested_by", "fulfilled_by").order_by("-created_at")
    serializer_class = MaterialTransferRequestSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        transfer = serializer.save(requested_by=self.request.user)
        material_title = transfer.material.title
        library = transfer.material.library
        _notify_role_users(
            roles=["STACK STAFF"],
            message=(
                f"Transfer request #{str(transfer.id)[:8]} created for "
                f"'{material_title}' ({transfer.requested_quantity} copies)."
            ),
            library=library,
        )

    @action(detail=True, methods=["post"], url_path="fulfill")
    def fulfill(self, request, pk=None):
        from django.db import transaction as db_transaction
        
        transfer = self.get_object()

        if transfer.status != "PENDING":
            return Response(
                {"detail": "Only pending transfers can be fulfilled."},
                status=status.HTTP_400_BAD_REQUEST
            )

        with db_transaction.atomic():
            material = PhysicalMaterial.objects.select_for_update().get(pk=transfer.material.pk)

            if material.available_copies < transfer.requested_quantity:
                return Response(
                    {"detail": "Not enough copies available in STACK."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 1. Decrease total_copies and available_copies from the source STACK material
            material.total_copies = max(0, material.total_copies - transfer.requested_quantity)
            material.available_copies = max(0, material.available_copies - transfer.requested_quantity)
            material.save()

            # 2. Look for an existing SHELF material with the same attributes in the same library
            shelf_material = PhysicalMaterial.objects.filter(
                title=material.title,
                author=material.author,
                library=material.library,
                location="SHELF"
            ).first()

            if shelf_material:
                shelf_material.total_copies += transfer.requested_quantity
                shelf_material.available_copies = (shelf_material.available_copies or 0) + transfer.requested_quantity
                shelf_material.save()
            else:
                shelf_material = PhysicalMaterial.objects.create(
                    title=material.title,
                    author=material.author,
                    category=material.category,
                    genre=material.genre,
                    published_date=material.published_date,
                    department=material.department,
                    language=material.language,
                    isbn=material.isbn,
                    price=material.price,
                    condition=material.condition,
                    location="SHELF",
                    total_copies=transfer.requested_quantity,
                    available_copies=transfer.requested_quantity,
                    can_borrow=material.can_borrow,
                    library=material.library,
                    created_by=request.user
                )

            # 3. Complete the transfer request
            transfer.status = "COMPLETED"
            transfer.fulfilled_by = request.user
            transfer.transferred_quantity = transfer.requested_quantity
            transfer.completed_at = timezone.now()
            transfer.save()

        material_title = material.title
        _notify_role_users(
            roles=["FRONT DESK STAFF"],
            message=(
                f"Transfer request fulfilled: '{material_title}' "
                f"({transfer.transferred_quantity} copies) moved to shelf."
            ),
            library=material.library,
        )
        if transfer.requested_by_id:
            from user_mgt.models import Notification

            Notification.objects.create(
                member_id=transfer.requested_by,
                message=(
                    f"Your transfer for '{material_title}' was fulfilled. "
                    f"{transfer.transferred_quantity} copies are now on the shelf."
                )[:200],
                status="UNREAD",
            )

        return Response(
            self.get_serializer(transfer).data
        )

    @action(detail=True, methods=["post"], url_path="cancel")
    def cancel(self, request, pk=None):
        transfer = self.get_object()
        if transfer.status != "PENDING":
            return Response({"detail": "Only pending transfers can be cancelled."}, status=status.HTTP_400_BAD_REQUEST)
        
        transfer.status = "CANCELLED"
        transfer.completed_at = timezone.now()
        transfer.save()

        return Response(self.get_serializer(transfer).data)


class MaterialBarcodeLookupAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        code = str(request.query_params.get("code") or "").strip()
        if not code:
            return Response({"detail": "Query parameter 'code' is required."}, status=status.HTTP_400_BAD_REQUEST)

        result = lookup_book_metadata(code)
        return Response(result, status=status.HTTP_200_OK)


import barcode
from barcode.writer import ImageWriter
from django.http import HttpResponse

class BarcodeImageAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk, *args, **kwargs):
        try:
            material = PhysicalMaterial.objects.get(pk=pk)
        except PhysicalMaterial.DoesNotExist:
            return Response({"detail": "Material not found."}, status=status.HTTP_404_NOT_FOUND)
        
        if not material.barcode:
            material.save()
            
        try:
            Code128 = barcode.get_barcode_class('code128')
            barcode_obj = Code128(material.barcode, writer=ImageWriter())
            from io import BytesIO
            buffer = BytesIO()
            barcode_obj.write(buffer)
            return HttpResponse(buffer.getvalue(), content_type="image/png")
        except Exception as e:
            return Response({"detail": f"Failed to generate barcode image: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


import openpyxl
from rest_framework.parsers import MultiPartParser, FormParser

class PhysicalMaterialXLSImportAPIView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        try:
            _get_staff_profile_or_error(request.user)
        except ValidationError as val_err:
            return Response(val_err.detail, status=status.HTTP_403_FORBIDDEN)
            
        actor_library = get_user_library(request.user)
        if not actor_library and not is_super_admin(request.user):
            return Response({"detail": "Your account is not assigned to a library yet."}, status=status.HTTP_400_BAD_REQUEST)

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        
        if not file_obj.name.endswith('.xlsx'):
            return Response({"detail": "Invalid file format. Only .xlsx files are supported."}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            sheet = wb.active
            
            headers = {}
            for col_idx in range(1, sheet.max_column + 1):
                val = sheet.cell(row=1, column=col_idx).value
                if val:
                    headers[str(val).strip().lower()] = col_idx
            
            required_fields = ['title', 'author', 'category', 'genre', 'published_date', 'total_copies', 'price']
            missing = [f for f in required_fields if f not in headers]
            if missing:
                return Response({
                    "detail": f"Missing required columns: {', '.join(missing)}. File must have headers: title, author, category, genre, published_date, total_copies, price, department, language, isbn, condition, location, can_borrow."
                }, status=status.HTTP_400_BAD_REQUEST)
                
            created_count = 0
            errors = []
            
            for row_idx in range(2, sheet.max_row + 1):
                title_val = sheet.cell(row=row_idx, column=headers['title']).value
                if not title_val:
                    continue
                    
                try:
                    title = str(title_val).strip()
                    author = str(sheet.cell(row=row_idx, column=headers['author']).value or '').strip()
                    category = str(sheet.cell(row=row_idx, column=headers['category']).value or '').strip().upper()
                    genre = str(sheet.cell(row=row_idx, column=headers['genre']).value or '').strip()
                    
                    pub_date_val = sheet.cell(row=row_idx, column=headers['published_date']).value
                    from datetime import datetime, date
                    if isinstance(pub_date_val, (datetime, date)):
                        published_date = pub_date_val
                    else:
                        pub_date_str = str(pub_date_val).strip()
                        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                            try:
                                published_date = datetime.strptime(pub_date_str, fmt).date()
                                break
                            except ValueError:
                                pass
                        else:
                            raise ValueError(f"Invalid date format: {pub_date_str}. Use YYYY-MM-DD.")
                        
                    total_copies_val = sheet.cell(row=row_idx, column=headers['total_copies']).value
                    if total_copies_val is None:
                        raise ValueError("total_copies is required.")
                    total_copies = int(total_copies_val)
                    
                    price_val = sheet.cell(row=row_idx, column=headers['price']).value
                    if price_val is None:
                        raise ValueError("price is required.")
                    from decimal import Decimal
                    price = Decimal(str(price_val))
                    
                    department = str(sheet.cell(row=row_idx, column=headers.get('department', 999)).value or 'General').strip()
                    language = str(sheet.cell(row=row_idx, column=headers.get('language', 999)).value or 'English').strip()
                    
                    isbn_val = sheet.cell(row=row_idx, column=headers.get('isbn', 999)).value
                    isbn = str(isbn_val).strip() if isbn_val else None
                    
                    condition = str(sheet.cell(row=row_idx, column=headers.get('condition', 999)).value or 'GOOD').strip().upper()
                    if condition not in ['NEW', 'GOOD', 'FAIR', 'DAMAGED']:
                        condition = 'GOOD'
                        
                    location = str(sheet.cell(row=row_idx, column=headers.get('location', 999)).value or 'STACK').strip().upper()
                    if location not in ['STACK', 'SHELF']:
                        location = 'STACK'
                        
                    can_borrow_val = sheet.cell(row=row_idx, column=headers.get('can_borrow', 999)).value
                    can_borrow = True
                    if can_borrow_val is not None:
                        can_borrow = str(can_borrow_val).strip().lower() in ['1', 'true', 'yes', 'y']
                        
                    library = actor_library
                    if is_super_admin(request.user) and 'library' in headers:
                        lib_name = str(sheet.cell(row=row_idx, column=headers['library']).value or '').strip()
                        from backend.models import Library
                        lib = Library.objects.filter(name__iexact=lib_name).first()
                        if lib:
                            library = lib
                            
                    if not library:
                        raise ValueError("Library not assigned/resolved.")
                        
                    PhysicalMaterial.objects.create(
                        title=title,
                        author=author,
                        category=category,
                        genre=genre,
                        published_date=published_date,
                        total_copies=total_copies,
                        available_copies=total_copies,
                        price=price,
                        department=department,
                        language=language,
                        isbn=isbn,
                        condition=condition,
                        location=location,
                        can_borrow=can_borrow,
                        library=library,
                        created_by=request.user
                    )
                    created_count += 1
                except Exception as row_err:
                    errors.append(f"Row {row_idx}: {str(row_err)}")
                    
            return Response({
                "success": True,
                "created_count": created_count,
                "failed_count": len(errors),
                "errors": errors
            }, status=status.HTTP_200_OK if not errors else status.HTTP_207_MULTI_STATUS)
            
        except Exception as e:
            return Response({"detail": f"Failed to parse XLS file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


